import openpyxl
from openpyxl import load_workbook
wb = openpyxl.load_workbook(r"D:\1.xlsx") #导入现场记录表
print(wb.sheetnames)
ws = wb['应用系统']
print(ws['B2'])
test1 = ws['B2'].value #提取B2的内容
print(test1)
wb = load_workbook(r"D:\myfile.xlsx") 
zq = wb ["增强表"]
yb = wb ["一般表"]
if test1 is not None:
    zq.cell(2,2,test1) 
if test1 is not None:
    yb.cell(2,4,test1)
wb.save("D:\myfile.xlsx") #保存新文件